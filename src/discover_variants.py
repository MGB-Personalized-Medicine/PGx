import vcf
import traceback
import logging
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font


create_genotyping_report_logger = logging.getLogger('PgxTyper.DiscoverVariants')

def _categorize_WGS_call(DP, QD, MQ, FS, AX, chr, sex):
	fail_reasons = []
	qc = 'PASS'

	#We may at some point wish to set different DP thresholds for Males and Females, but this functionality has not been thoroughly tested yet
	if chr == 'X' and sex == 'M' and DP < 10:
		fail_reasons.append('DP')
	elif chr == 'X' and sex == 'F' and DP < 10:
		fail_reasons.append('DP')
	if QD < 4 and MQ < 60:
		fail_reasons.append('QD/MQ')
	if AX < 0.15:
		fail_reasons.append('AX')
	if FS > 35:
		fail_reasons.append('FS')

	if len(fail_reasons) > 0:
		qc = '|'.join(fail_reasons)

	return qc

def _categorize_WES_call(DP, QD, MQ, FS, AX, chr, sex):
	fail_reasons = []
	qc = 'PASS'

	#We may at some point wish to set different DP thresholds for Males and Females, but this functionality has not been thoroughly tested yet
	if chr == 'X' and sex == 'M' and DP < 10:
		fail_reasons.append('DP')
	elif chr == 'X' and sex == 'F' and DP < 10:
		fail_reasons.append('DP')
	if QD < 2:
		fail_reasons.append('QD')
	if AX < 0.2:
		fail_reasons.append('AX')
	if FS > 40:
		fail_reasons.append('FS')
		
	if len(fail_reasons) > 0:
		qc = '|'.join(fail_reasons)

	return qc

#Given a VCF, pull in all variant-specific annotations for PGx sites; flag FPS, multiallelic sites, sites within variant calling ROI that aren't actual PGx sites
#Write output as .txt and formatted .xlsx; .txt is used by later steps, .xlsx is for easier manual review
def _annotate_genotypes(genotyped_vcf_fname,annotated_bed_fname,sex,genotyping_report_txt,genotyping_report_xlsx,thresholds):
	roi_annotations = {}
	with open(annotated_bed_fname, 'r') as f:
		for line in f:
			chr, start, end, ref, alt, variant_type, gene, transcript, region, cms_nomenclature, cdna_change, aa_change, variant_id, variant_loci = line.rstrip().split('\t')
			if variant_loci not in roi_annotations:
				roi_annotations[variant_loci] = {}
			roi_annotations[variant_loci][ref + '-' + alt] = {'REF' : ref, 'ALT' : set(alt.split("/")), 'VARIANT_ID' : variant_id, 'GENE' : gene, 'TRANSCRIPT' : transcript, 'REGION' : region, 'CMS_NOMENCLATURE' : cms_nomenclature, 'CDNA_CHANGE' : cdna_change, 'AA_CHANGE' : aa_change}

	wb = openpyxl.Workbook()
	ws = wb.active
	with open(genotyped_vcf_fname,'r') as genotyped_vcf:
		genotyped_vcf = vcf.Reader(open(genotyped_vcf_fname,'r'))
		with open(genotyping_report_txt, 'w') as out_f:
			output_header = ['CHR','POS','VARIANT_ID','GT','GENOTYPE', 'DP','REF', 'ALT', 'AX', 'QD','FS','MQ', 'GQ', 'ZYGOSITY', 'GENE', 'TRANSCRIPT','REGION','CMS_NOMENCLATURE','CDNA_CHANGE', 'AA_CHANGE']
			out_f.write("\t".join(output_header) + '\n')
			row = 1
			for column in range(0, len(output_header)):
				ws.cell(row=row, column=column+1).value = output_header[column]
				ws.cell(row=row, column=column+1).font = Font(bold=True)

			#Process VCF records for variant discovery
			for record in genotyped_vcf:
				variant_loci = str(record.CHROM)+":"+str(record.POS)
				variant_nucleotides = [record.REF,set(str(altallele) for altallele in record.ALT)]
			    
				assert(len(record.samples)==1)
				sample = record.samples[0]
			    
			    ##Obtain Variant calling metrics
				try:
					QD = record.INFO['QD']
				except KeyError:
					QD = "NA"
				try:
					FS = record.INFO['FS']
				except KeyError:
					FS = "NA"
				try: 
					MQ = record.INFO['MQ']
				except KeyError:
					MQ = "NA"
				try:
					GQ = sample['GQ']
				except KeyError:
					GQ = "NA"
				try:
					DP = sample['DP']
				except:
					DP = "NA"
				AX = {} #compute allelic-fraction for each alternative allele
				try:
					all_allele_depths = sample['AD']
					total_depth = sum(all_allele_depths)
					AD = {}
					for alt_allele,alt_allele_depth in zip(record.ALT,all_allele_depths[1:]):
						AX[str(alt_allele)] = [(float(alt_allele_depth)/float(total_depth))] 
						AD[str(alt_allele)] = alt_allele_depth
				except (AttributeError,ZeroDivisionError) as az: ## (0/0 does not have AD, but has DP) <<OR>> if AD = [0,0,0] and DP = 0 
					AX['None'] = [0] 

				qc_metrics = [DP,QD,FS,MQ,AX,GQ]
			    
			    ## Evaluate variant for PASS/FAIL criteria based on variant calling metrics
				if sample['GT'] == './.':
					zygosity = 'NA'
					qc_eval = 'NA'
					GT = 'NoCall'
					AX['None'].append(zygosity) 
			    
				elif sample['GT'] == '0/1' and record.ALT == [None]:
					#Weird handling; this is an indel that has another row in the VCF representing it
					continue

				elif sample['GT'] == '0/0' or record.ALT == [None]:
					zygosity = 'WT'
					GT = "/".join([record.REF,record.REF])
					if record.CHROM == 'X' and sex == 'M': 	#We may at some point wish to set different DP thresholds for Males and Females, but this functionality has not been thoroughly tested yet
						if DP >= 10:
							qc_eval = 'PASS'
					elif DP >= 10: 
						qc_eval = 'PASS'
					else:
						qc_eval = 'DP'
					AX['None'].append(zygosity)
			    


				else: ## Any other genotype (0/1, 1/1, 1/2 (or any other multi-allelic))
					try:
						if len(AX) == 1: #Not multiallelic
							if thresholds == 'WGS':
								qc_eval = _categorize_WGS_call(DP, QD, MQ, FS, list(AX.values())[0][0], record.CHROM, sex)
							elif thresholds == 'WES':
								qc_eval = _categorize_WES_call(DP, QD, MQ, FS, list(AX.values())[0][0], record.CHROM, sex)
							else:
								qc_eval = 'NOEVAL'
						else:
							qc_eval = 'MULTIALLELIC'
					
					except Exception as _AnyMissingMetricsException: 
						create_genotyping_report_logger.error("Flagged QC Metrics: Chr:{} Pos:{} Metrics: {}".format(record.CHROM,record.POS,qc_metrics))
						create_genotyping_report_logger.error("Marking this variant as 'FAIL'(QC). Follow up required.")
						create_genotyping_report_logger.error("ERROR: {}".format(_AnyMissingMetricsException))
						qc_eval = 'FAIL'
			        
			        ## Assign zygosity based on allelic-fraction 
					for alts in AX.keys():
						if qc_eval != 'PASS':
							zygosity = 'NA=' + qc_eval
						elif AX[alts][0] <= 0.1:
							zygosity = 'WT'
						elif 0.1 < AX[alts][0] <= 0.25:
							zygosity = 'HETWT'
						elif 0.25 < AX[alts][0] <= 0.75:
							zygosity = 'HET'
						elif 0.75 < AX[alts][0] <= 0.9:
							zygosity = 'HOMHET'
						elif AX[alts][0] > 0.9:
							zygosity = 'HOM'
						else: 
							zygosity = 'NA'           
						if len(AX) > 1:
							zygosity = 'MULTIALLELIC'
						AX[alts].append(zygosity)
			        
			        ## Describe genotypes with representative nucleotides
					all_nucleotides = [record.REF]
					for a in record.ALT:
						all_nucleotides.append(str(a))
					GT = all_nucleotides[int(sample['GT'].split("/")[0])]+"/"+all_nucleotides[int(sample['GT'].split("/")[1])]
			        
				qc_metrics.append(qc_eval)
				
				#Prepare output
				report_variant = {}
				report_variant['CHR'] = record.CHROM
				report_variant['POS'] = record.POS
				report_variant['GENOTYPE'] = GT
				report_variant['GT'] = sample['GT']
				report_variant['QD'] = QD
				report_variant['REF'] = all_allele_depths[0]
				report_variant['DP'] = DP
				report_variant['FS'] = FS
				report_variant['MQ'] = MQ
				report_variant['GQ'] = GQ
				report_variant['QC_EVAL'] = qc_eval
				report_variant['ALT'] = 'NA'
				report_variant['AX'] = 'NA'

				variant_nucleotides[1].discard('None')
				alt_alleles = variant_nucleotides[1]
				for alt in alt_alleles:
					if report_variant['ALT'] != 'NA':
						report_variant['ALT'] += '|' + alt + ':' + str(AD[alt])
						report_variant['AX'] += '|' + alt + ':' + str(AX[alt][0])
					else:
						report_variant['ALT'] = alt + ':' + str(AD[alt])
						report_variant['AX'] = alt + ':' + str(AX[alt][0])

				report_variant['VARIANT_ID'] = 'NOT_IN_ROI'
				report_variant['GENE'] = 'NOT_IN_ROI'
				report_variant['TRANSCRIPT'] = 'NOT_IN_ROI'
				report_variant['REGION'] = 'NOT_IN_ROI'
				report_variant['CMS_NOMENCLATURE'] = 'NOT_IN_ROI'
				report_variant['CDNA_CHANGE']  = 'NOT_IN_ROI'
				report_variant['AA_CHANGE'] = 'NOT_IN_ROI'
				report_variant['ZYGOSITY'] = 'NOT_IN_ROI'


				if variant_loci in roi_annotations.keys(): #This is a position we care about
					this_variant_annotations = roi_annotations[variant_loci][list(roi_annotations[variant_loci])[0]]
					if len(roi_annotations[variant_loci]) > 1:
						if ref + '-' + alt in roi_annotations[variant_loci]: #Try to get the right annotations if this is a variant we have annotations for
							this_variant_annotations = roi_annotations[variant_loci][ref + '-' + alt]
					row += 1
					report_variant['VARIANT_ID'] = this_variant_annotations['VARIANT_ID']
					report_variant['GENE'] = this_variant_annotations['GENE']
					report_variant['TRANSCRIPT'] = this_variant_annotations['TRANSCRIPT']
					report_variant['REGION'] = this_variant_annotations['REGION']
					report_variant['CMS_NOMENCLATURE'] = this_variant_annotations['CMS_NOMENCLATURE']
					report_variant['CDNA_CHANGE']  = this_variant_annotations['CDNA_CHANGE']
					report_variant['AA_CHANGE'] = this_variant_annotations['AA_CHANGE']
					found_alt_in_roi = list(variant_nucleotides[1].intersection(this_variant_annotations['ALT']))
					
					if report_variant['GT'] == '0/0':
						if qc_eval != 'PASS':
							report_variant['ZYGOSITY'] = 'NA=' + qc_eval
						else:
							report_variant['ZYGOSITY'] = 'WT'
					elif len(alt_alleles) > 1:
						report_variant['ZYGOSITY'] = 'MULTIALLELIC'
					elif len(found_alt_in_roi) >= 1:

						report_variant['ZYGOSITY'] = found_alt_in_roi[0]+":"+str(AX[found_alt_in_roi[0]][1])
					else:
						report_variant['ZYGOSITY'] = 'ALT'

					#Special handling for F5
					if 'F5' in report_variant['GENE']:
						if report_variant['GENOTYPE'] == 'C/C':
							report_variant['ZYGOSITY'] = 'WT'
						elif report_variant['GENOTYPE'] == 'T/T':
							report_variant['ZYGOSITY'] = 'HOM'
						elif 'T/' in report_variant['GENOTYPE'] or '/T' in report_variant['GENOTYPE']:
							report_variant['ZYGOSITY'] = 'HET'

					out_f.write('\t'.join([str(report_variant[header]) for header in output_header]) + '\n')
					for column in range(0,len(output_header)):
						ws.cell(row=row, column=column+1).value = report_variant[output_header[column]]
						if report_variant['ZYGOSITY'] != 'WT':
							if report_variant['ZYGOSITY'] in ['MULTIALLELIC', 'ALT'] or 'NA' in report_variant['ZYGOSITY']:
								ws.cell(row=row, column=column+1).font = Font(italic=True, color="fc0808", bold=True)
							else:
								ws.cell(row=row, column=column+1).fill = PatternFill(start_color="feedc6",end_color="feedc6",fill_type="solid")

				else: #this is a position we don't care about, unless it's non-WT, in which case output it for manual review
					if len(alt_alleles) == 0:
						pass
					else:
						row += 1
						if len(alt_alleles) > 1:
							report_variant['ZYGOSITY'] = 'NotInROI|MULTIALLELIC'
							ws.cell(row=row, column=column+1).font = Font(italic=True, color="fc0808", bold=True)
						elif len(alt_alleles) == 1:
							report_variant['ZYGOSITY'] = 'NOT_IN_ROI'
						out_f.write('\t'.join([str(report_variant[header]) for header in output_header]) + '\n')
						for column in range(0,len(output_header)):
							ws.cell(row=row, column=column+1).value = report_variant[output_header[column]]
							ws.cell(row=row, column=column+1).font = Font(italic=True, color="fc0808", bold=True)

	wb.save(genotyping_report_xlsx)


