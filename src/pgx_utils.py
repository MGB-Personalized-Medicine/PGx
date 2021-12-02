import logging
import os
import sys
import collections
import openpyxl
from copy import copy

pgx_utils_logger = logging.getLogger('PgxTyper.DiscoverDiplotypes')

#Given variant calls and their annotations (created by discover_variants.py), identify associated diplotypes, use diplotypes as key to get associated notes and phenotypes
def _diplotypes_to_phenotypes(genotyping_report, gene_annotations_dir, genes_without_annotations, cyp2d6_gt_to_pheno_dir, sample):
    gene_phenotypes = {}
    all_genes = [ f.name for f in os.scandir(gene_annotations_dir) if f.is_dir() ]
    called_variant_zygosities = _get_zygosities_from_genotyping_report(genotyping_report)
    for gene in all_genes:
    # Using the gt_to_pheno file created form the Cyrius output file
        if gene == 'CYP2D6':
            gt2pheno_dir = os.path.abspath(cyp2d6_gt_to_pheno_dir)
            genotypes_phenotypes_fname = gt2pheno_dir + '/' + gene + '_gt_to_pheno.txt'
        else :
            genotypes_phenotypes_fname = gene_annotations_dir + '/' + gene + '/' + gene + '_gt_to_pheno.txt'

        with open(genotypes_phenotypes_fname, 'r') as f:
            header_fields = next(f).rstrip().split('\t')
            expected_headers = ['Diplotype', 'Phenotype', 'Diplotype Notes'] #These fields should always be present; additionally, each gene should have a set of variants associated with it
            if not(all(expected_header in  header_fields for expected_header in expected_headers)):
                pgx_utils_logger.error("Header Entry Missing from " + gene + " file. Failed to parse. Exiting.")
                sys.exit(1)

            variant_ids_to_match = [header_fields[variant_position] for variant_position in range(len(expected_headers),len(header_fields))]
            for line in f:
                fields = line.rstrip('\n').split('\t')
                zygosities = fields[len(expected_headers):len(header_fields)]
                # Special handling for CYP2D6, since the gt_to_pheno file contains sample specific pheno data and the bed file does not contain regions for CYP2D6
                if gene == 'CYP2D6' and fields[header_fields.index('Sample')] == sample:
                    gene_phenotypes[gene] = {}
                    gene_phenotypes[gene]['Diplotype'] = fields[header_fields.index('Diplotype')]
                    gene_phenotypes[gene]['Phenotype'] = fields[header_fields.index('Phenotype')]
                    gene_phenotypes[gene]['Diplotype Notes'] = fields[header_fields.index('Diplotype Notes')]
                elif dict(zip(variant_ids_to_match, zygosities)).items() <= called_variant_zygosities.items():
                    gene_phenotypes[gene] = {}
                    gene_phenotypes[gene]['Diplotype'] = fields[header_fields.index('Diplotype')]
                    gene_phenotypes[gene]['Phenotype'] = fields[header_fields.index('Phenotype')]
                    gene_phenotypes[gene]['Diplotype Notes'] = fields[header_fields.index('Diplotype Notes')]
        if gene not in gene_phenotypes: #Should usually indicate a variant called but failing QC, but could also occur if annotation tables are out of synch
            gene_phenotypes[gene] = {}
            gene_phenotypes[gene]['Diplotype'] = 'NA'
            gene_phenotypes[gene]['Phenotype'] = 'NA'

    #Special handling for G6PD; we don't yet have annotation tables constructed (lots of combinations, sex-specific), but still want to (quickly) get it incorporated into the FDA outputs; this is poor design, but works for now...
    extra_genes_variants = {}
    with open(genes_without_annotations, 'r') as f_extra_genes:
        for line in f_extra_genes:
            gene, drugs = line.rstrip().split('\t')
            extra_genes_variants[drugs] = {}
            extra_genes_variants[drugs][gene] = []
            with open(genotyping_report, 'r') as f_genotyping_report:
                header_fields = next(f_genotyping_report).rstrip('\n').split('\t')
                for line in f_genotyping_report:
                    fields = line.rstrip('\n').split('\t')
                    this_gene = fields[header_fields.index('GENE')]
                    variant_id = fields[header_fields.index('VARIANT_ID')]
                    cms_nomenclature = fields[header_fields.index('CMS_NOMENCLATURE')]
                    zygosity = _format_zygosity(fields[header_fields.index('ZYGOSITY')])
                    if this_gene == gene and zygosity != "WT":
                        extra_genes_variants[drugs][gene].append(cms_nomenclature + '=' + zygosity)

    return gene_phenotypes, extra_genes_variants

#Get CPIC drug info given gene (and gene group) phenotypes
def _phenotypes_to_CPIC_drug_info(gene_phenotypes, gene_annotations_dir):
    phenotypes_CPIC_drug_info = {}
    drugs_per_gene = {}
    for gene in gene_phenotypes:
        phenotypes_dosages_fname = gene_annotations_dir + '/' + gene + '/' + gene + '_pheno_to_dose.txt'
        phenotypes_CPIC_drug_info[gene] = []
        drugs_per_gene[gene] = set()

        with open(phenotypes_dosages_fname, 'r') as f:
            header_fields = next(f).rstrip().split('\t')
            for line in f:
                fields = line.rstrip('\n').split('\t')
                drugs_per_gene[gene].add(fields[header_fields.index('Drug')])
                if fields[header_fields.index('Phenotype')] == gene_phenotypes[gene]['Phenotype']:
                    drug_annotations = {}
                    drug_annotations['Drug'] = fields[header_fields.index('Drug')]
                    drug_annotations['Drug Class or Action'] = fields[header_fields.index('Drug Class or Action')]
                    drug_annotations['Phenotype'] = fields[header_fields.index('Phenotype')]
                    drug_annotations['Dosage Information'] = fields[header_fields.index('Dosage Information')]
                    drug_annotations['Category'] = fields[header_fields.index('Category')]
                    drug_annotations['Interpretation'] = fields[header_fields.index('Interpretation')]
                    drug_annotations['Notes'] = fields[header_fields.index('Notes')]
                    drug_annotations['References'] = fields[header_fields.index('References')]
                    phenotypes_CPIC_drug_info[gene].append(drug_annotations)
    return phenotypes_CPIC_drug_info, drugs_per_gene

def _write_CPIC_output(gene_phenotypes, gene_drug_pgx_annotations, drugs_per_gene, pgx_report_nonfda_template_fname, pgx_report_cpic_fname, accession):
    standard_dosage_drugs = {}
    atypical_dosage_drugs = {}
    genes_without_diplotypes = []
    num_atypical = 0
    num_standard = 0
    for gene, drugs in gene_drug_pgx_annotations.items():
        for pgx_drug_annotations in drugs:
            if pgx_drug_annotations['Phenotype'] != 'CYP2D6 Indeterminate': # Special handling for 'CYP2D6 Indeterminate' as those go into a separate table in the report
                if pgx_drug_annotations['Category'] == 'STANDARD' or pgx_drug_annotations['Category'] == '':
                    if gene not in standard_dosage_drugs:
                        standard_dosage_drugs[gene] = []
                    standard_dosage_drugs[gene].append(pgx_drug_annotations)
                    num_standard += 1
                else:
                    if gene not in atypical_dosage_drugs:
                        atypical_dosage_drugs[gene] = []
                    atypical_dosage_drugs[gene].append(pgx_drug_annotations)
                    num_atypical += 1

    wb = openpyxl.load_workbook(pgx_report_nonfda_template_fname)
    ws = wb.active
    template_merged_cells = ws.merged_cell_ranges[:]

    block_height = 7
    #Copy template to make space for all atypical dosages
    for i in range(0, num_atypical):
        for column in ['A', 'B', 'C', 'D']:
            for row in range(2,7) :
                ws.row_dimensions[row + block_height * i].height = ws.row_dimensions[row].height
                _copy_cell(ws[column + str(row)], column + str(row + block_height * i), ws)
        for template_merged_cell in template_merged_cells:
            start, end = template_merged_cell.split(':')
            start_column = start[0]
            start_row = start[1:]
            end_column = end[0]
            end_row = end[1:]
            new_merged_cell = start_column + str(int(start_row) + block_height * i) + ':' + end_column + str(int(end_row) + block_height * i)
            ws.merge_cells(new_merged_cell)

    current_atypical_variant_count = 0
    ws.cell('A1').value = accession
    for gene in atypical_dosage_drugs:
        for gene_drug_annotations in atypical_dosage_drugs[gene]:
            ws.cell('A' + str(3 + current_atypical_variant_count * block_height)).value = gene_drug_annotations['Drug'].strip('"')
            ws.cell('B' + str(3 + current_atypical_variant_count * block_height)).value = gene_drug_annotations['Dosage Information'].strip('"')
            ws.cell('C' + str(3 + current_atypical_variant_count * block_height)).value = gene_phenotypes[gene]['Diplotype'].strip('"')
            ws.cell('D' + str(3 + current_atypical_variant_count * block_height)).value = gene_drug_annotations['Phenotype'].strip('"')
            if gene_phenotypes[gene]['Diplotype Notes'].rstrip(' ') != '':
                ws.cell('A' + str(4 + current_atypical_variant_count * block_height)).value = 'INTERPRETATION: ' + gene_drug_annotations['Interpretation'].strip('"') + '\n' + gene_phenotypes[gene]['Diplotype Notes'].rstrip(' ')
            else:
                ws.cell('A' + str(4 + current_atypical_variant_count * block_height)).value = 'INTERPRETATION: ' + gene_drug_annotations['Interpretation'].strip('"')
            ws.cell('A' + str(5 + current_atypical_variant_count * block_height)).value = 'NOTES: ' + gene_drug_annotations['Notes'].strip('"')
            ws.cell('A' + str(6 + current_atypical_variant_count * block_height)).value = 'REFERENCES: ' + gene_drug_annotations['References'].strip('"')
            current_atypical_variant_count += 1

    row = 6 + (current_atypical_variant_count - 1) * block_height + 6 #Offset from atypical section
    ws.row_dimensions[row].height = ws.row_dimensions[1].height
    _copy_cell(ws['A2'],'A' + str(row),ws)
    _copy_cell(ws['B2'],'B' + str(row),ws)
    _copy_cell(ws['C2'],'C' + str(row),ws)
    _copy_cell(ws['D2'],'D' + str(row),ws)

    ws.cell('A' + str(row)).value = 'Drugs for which standard dosage is recommended'
    ws.merge_cells('A' + str(row) + ':' + 'D' + str(row))
    row += 1
    ws.row_dimensions[row].height = ws.row_dimensions[1].height
    _copy_cell(ws['A2'],'A' + str(row),ws)
    _copy_cell(ws['B2'],'B' + str(row),ws)
    _copy_cell(ws['C2'],'C' + str(row),ws)
    _copy_cell(ws['D2'],'D' + str(row),ws)

    ws.cell('A' + str(row)).value = 'Drug'
    ws.cell('B' + str(row)).value = 'Drug Class or Action'
    ws.cell('C' + str(row)).value = 'Diplotype'
    ws.cell('D' + str(row)).value = 'Phenotype'
    row += 1

    for gene in standard_dosage_drugs:
        for gene_drug_annotations in standard_dosage_drugs[gene]:
            ws.row_dimensions[row].height = ws.row_dimensions[2].height
            _copy_cell(ws['A3'],'A' + str(row),ws)
            _copy_cell(ws['B3'],'B' + str(row),ws)
            _copy_cell(ws['C3'],'C' + str(row),ws)
            _copy_cell(ws['D3'],'D' + str(row),ws)
            ws.cell('A' + str(row)).value = gene_drug_annotations['Drug'].strip('"')
            ws.cell('B' + str(row)).value = gene_drug_annotations['Drug Class or Action'].strip('"')
            ws.cell('C' + str(row)).value = gene_phenotypes[gene]['Diplotype'].strip('"')
            ws.cell('D' + str(row)).value = gene_drug_annotations['Phenotype'].strip('"')
            row += 1

    #Genes for which phenotypes could not be determined (Currently, CYP2D6 indeterminate variants to be placed here)
    row += 4
    ws.row_dimensions[row].height = ws.row_dimensions[1].height
    _copy_cell(ws['A2'],'A' + str(row),ws)
    _copy_cell(ws['B2'],'B' + str(row),ws)
    _copy_cell(ws['C2'],'C' + str(row),ws)
    _copy_cell(ws['D2'],'D' + str(row),ws)

    ws.cell('A' + str(row)).value = 'Genes for which phenotypes could not be determined - the patient carries at least one uncertain and/or unknown function allele.'
    ws.merge_cells('A' + str(row) + ':' + 'D' + str(row))

    row += 1
    ws.row_dimensions[row].height = ws.row_dimensions[1].height
    _copy_cell(ws['A2'], 'A' + str(row), ws)
    ws.cell('A' + str(row)).value = 'Gene'
    _copy_cell(ws['B2'], 'B' + str(row), ws)
    ws.cell('B' + str(row)).value = 'Drug'
    _copy_cell(ws['C2'],'C' + str(row),ws)
    ws.cell('C' + str(row)).value = 'Diplotype'
    _copy_cell(ws['D2'],'D' + str(row),ws)
    ws.cell('D' + str(row)).value = 'Phenotype'

    for gene in gene_phenotypes:
        if gene_phenotypes[gene]['Phenotype'] == 'CYP2D6 Indeterminate':
            for drug in drugs_per_gene[gene]:
                row += 1
                ws.row_dimensions[row].height = ws.row_dimensions[1].height
                _copy_cell(ws['A3'], 'A' + str(row), ws)
                _copy_cell(ws['B3'], 'B' + str(row), ws)
                _copy_cell(ws['C3'],'C' + str(row),ws)
                _copy_cell(ws['D3'],'D' + str(row),ws)
                ws.cell('A' + str(row)).value = gene
                ws.cell('B' + str(row)).value = drug
                ws.cell('C' + str(row)).value = gene_phenotypes[gene]['Diplotype'].strip('"')
                ws.cell('D' + str(row)).value = gene_phenotypes[gene]['Phenotype'].strip('"')

    #Genes for which diplotypes could not be determined (typically, this means we identified some FP variant calls)
    row += 4
    ws.row_dimensions[row].height = ws.row_dimensions[1].height
    _copy_cell(ws['A2'],'A' + str(row),ws)
    _copy_cell(ws['B2'],'B' + str(row),ws)

    ws.cell('A' + str(row)).value = 'Genes for which diplotypes could not be determined'
    ws.merge_cells('A' + str(row) + ':' + 'B' + str(row))

    row += 1
    ws.row_dimensions[row].height = ws.row_dimensions[1].height
    _copy_cell(ws['A2'], 'A' + str(row), ws)
    ws.cell('A' + str(row)).value = 'Gene'
    _copy_cell(ws['B2'], 'B' + str(row), ws)
    ws.cell('B' + str(row)).value = 'Drug'

    for gene in gene_phenotypes:
        if gene_phenotypes[gene]['Diplotype'] == 'NA':
            for drug in drugs_per_gene[gene]:
                row += 1
                ws.row_dimensions[row].height = ws.row_dimensions[1].height
                _copy_cell(ws['A3'], 'A' + str(row), ws)
                _copy_cell(ws['B3'], 'B' + str(row), ws)
                ws.cell('A' + str(row)).value =    gene
                ws.cell('B' + str(row)).value =    drug

    wb.save(pgx_report_cpic_fname)

#List of FDA drugs associated with each gene
def _read_FDA_drug_file(fname):
    druggroups = {}
    with open(fname, 'r') as f:
        next(f)
        for line in f:
            gene_group, drugs = line.rstrip('\n').split('\t')
            gene_group = gene_group.replace('/','_')
            if drugs not in druggroups:
                druggroups[drugs] = []
            druggroups[drugs].append(gene_group)
    return druggroups

def _gather_FDA_output(gene_drug_pgx_annotations, druggroups):
    annotations_to_write = ['Diplotype', 'Metabolism' , 'FDA-Drugs with PGx Labelling'] ##Include GeneGroup also, by default
    outputs = {}
    for druggroup in druggroups:
        for genegroup in druggroups[druggroup]:
            if genegroup in gene_drug_pgx_annotations:
                diplotype = gene_drug_pgx_annotations[genegroup]['Diplotype']
                metabolism = gene_drug_pgx_annotations[genegroup]['Phenotype']
                if diplotype == 'NA':
                    diplotype = genegroup + ' genotype could not be determined'
                    metabolism = genegroup + ' phenotype could not be determined'
                if 'Diplotype Notes' in gene_drug_pgx_annotations[genegroup]:
                    diplotype_note = gene_drug_pgx_annotations[genegroup]['Diplotype Notes']
                else:
                    diplotype_note = ''
                if druggroup not in outputs:
                    outputs[druggroup] = []
                outputs[druggroup].append([genegroup,diplotype,metabolism,diplotype_note])
            else:
                #Check to see if this is a single gene that's part of a larger "gene group"; currently this only applies to CYP2C9, VKORC1, and CYP4F2
                gene = genegroup
                for genegroup_check in gene_drug_pgx_annotations:
                    if gene in genegroup_check:
                        diplotypes_in_group = gene_drug_pgx_annotations[genegroup_check]['Diplotype'].split(';')
                        this_gene_index = 0
                        gene_index_count = 0
                        for diplotype in diplotypes_in_group:
                            if gene in diplotype:
                                this_gene_index = gene_index_count
                            gene_index_count += 1
                        diplotype = diplotypes_in_group[this_gene_index].lstrip(' ')
                        metabolism = gene_drug_pgx_annotations[genegroup_check]['Phenotype'].split('&')[this_gene_index].lstrip(' ')
                        if 'Diplotype Notes' in gene_drug_pgx_annotations[genegroup_check]:
                            diplotype_note = gene_drug_pgx_annotations[genegroup_check]['Diplotype Notes']
                        else:
                            diplotype_note = ''
                        if druggroup not in outputs:
                            outputs[druggroup] = []
                        outputs[druggroup].append([genegroup,diplotype,metabolism,diplotype_note])
                        break
    return outputs

def _write_FDA_output(pgx_report_fda_template_fname, fda_drugs_to_output, genes_variants_without_annotation, pgx_report_fda_fname, accession):
    diplotype_notes_symbols = ['i','ii','iii','iv','v', 'vi', 'vii', 'viii', 'ix', 'x']
    diplotype_note_count = 0
    all_diplotype_notes = []
    wb = openpyxl.load_workbook(pgx_report_fda_template_fname)
    ws = wb.active

    ws.cell('A1').value = accession
    current_row = 3 #header is taken care of by the template
    for druggroup in fda_drugs_to_output:
        genegroups = [gene[0] for gene in fda_drugs_to_output[druggroup]] #Don't need this?
        genotypes = [gene[1] for gene in fda_drugs_to_output[druggroup]]
        phenotypes = [gene[2] for gene in fda_drugs_to_output[druggroup]]
        diplotype_notes = [gene[3] for gene in fda_drugs_to_output[druggroup]]
        #Append diplotype note to genotypes if applicable
        for i in range(0, len(genotypes)):
            if diplotype_notes[i] != "":
                genotypes[i] = genotypes[i] + " (" + diplotype_notes_symbols[diplotype_note_count] + ")"
                diplotype_note_count += 1
                all_diplotype_notes.append(diplotype_notes[i])
        if druggroup != 'N/A':
            ws.cell('A' + str(current_row)).value = '\n'.join(genotypes)
            ws.cell('B' + str(current_row)).value = '\n'.join(phenotypes)
            ws.cell('C' + str(current_row)).value = druggroup
            current_row += 1
        else: #Separate row for each gene with 'N/A' drug (no FDA drugs)
            for gene in range(0, len(genotypes)):
                ws.cell('A' + str(current_row)).value = genotypes[gene]
                ws.cell('B' + str(current_row)).value = phenotypes[gene]
                ws.cell('C' + str(current_row)).value = druggroup
                current_row += 1


    for drugs in genes_variants_without_annotation:
        ws.row_dimensions[current_row].height = ws.row_dimensions[current_row-1].height
        _copy_cell(ws['A' + str(current_row-1)], 'A' + str(current_row), ws)
        ws.cell('A' + str(current_row)).value = ''
        _copy_cell(ws['B' + str(current_row-1)], 'B' + str(current_row), ws)
        ws.cell('B' + str(current_row)).value = 'N/A'
        _copy_cell(ws['C' + str(current_row-1)], 'C' + str(current_row), ws)
        ws.cell('C' + str(current_row)).value = drugs

        for gene in genes_variants_without_annotation[drugs]:
            if ws.cell('A' + str(current_row)).value != '':
                ws.cell('A' + str(current_row)).value += '\n'

            if len(genes_variants_without_annotation[drugs][gene]) >= 1:
                ws.cell('A' + str(current_row)).value += gene + ' variant(s): ' +  '|'.join(genes_variants_without_annotation[drugs][gene])
            else:
                ws.cell('A' + str(current_row)).value += gene + ' wild type at all tested positions'
        current_row += 1

    #Add diplotype note footers
    if len(all_diplotype_notes) > 0:
        new_merged_cell = 'A' + str(current_row) + ':' + 'C' + str(current_row)
        ws.merge_cells(new_merged_cell)
        _copy_cell(ws['A' + str(current_row-1)], 'A' + str(current_row), ws)
        _copy_cell(ws['B' + str(current_row-1)], 'B' + str(current_row), ws)
        _copy_cell(ws['C' + str(current_row-1)], 'C' + str(current_row), ws)
        ws.cell('A' + str(current_row)).value = ""
        ws.cell('B' + str(current_row)).value = ""
        ws.cell('C' + str(current_row)).value = ""
        #ws.cell('A' + str(current_row)).border = border_style
        #ws.cell('B' + str(current_row)).border = border_style
        #ws.cell('C' + str(current_row)).border = border_style
        for i in range(0, len(all_diplotype_notes)):
            if ws.cell('A' + str(current_row)).value == "":
                ws.cell('A' + str(current_row)).value = '(' + diplotype_notes_symbols[i] + ') '+ all_diplotype_notes[i]
            else:
                ws.cell('A' + str(current_row)).value = ws.cell('A' + str(current_row)).value + '\n(' + diplotype_notes_symbols[i] + ') '+ all_diplotype_notes[i]

    wb.save(pgx_report_fda_fname)

def _get_zygosities_from_genotyping_report(fname):
    called_variant_zygosities = {}
    with open(fname, 'r') as f:
        header_fields = next(f).rstrip('\n').split('\t')
        for line in f:
            fields = line.rstrip('\n').split('\t')
            variant_id = fields[header_fields.index('VARIANT_ID')]
            zygosity = _format_zygosity(fields[header_fields.index('ZYGOSITY')])
            called_variant_zygosities[variant_id] = zygosity
    return called_variant_zygosities

def _format_zygosity(zyg):
    if ":" in zyg:
        zyg = zyg.split(":")[1]
    else:
        pass
    return zyg

def _copy_cell(source_cell, coord, tgt):
    tgt[coord].value = source_cell.value
    if source_cell.has_style:
        tgt[coord]._style = source_cell._style
    return tgt[coord]
